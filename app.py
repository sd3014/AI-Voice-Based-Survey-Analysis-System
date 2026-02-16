from flask import Flask, render_template, request, jsonify
import os
import re
import requests
from docx import Document
import openpyxl
from openpyxl.styles import Alignment
from fuzzywuzzy import fuzz

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

questions = []
survey_topic = ""
user_responses = []
explanation_cache = {}

def get_ai_reply(prompt):
    try:
        response = requests.post(
            "http://localhost:11434/api/generate",
            json={
                "model": "mistral",  # You can also try "llama3" or "mistral"
                "prompt": prompt,
                "stream": False
            }
        )
        raw = response.json()["response"].strip()

        # 🧹 Remove hallucinated thinking or unnecessary meta-reasoning
        clean = re.sub(r"<think>.*?</think>", "", raw, flags=re.DOTALL).strip()
        return clean
    except Exception as e:
        return "Sorry, I couldn't generate a response right now."

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    global questions, survey_topic, user_responses, explanation_cache
    docx_file = request.files['docx']
    survey_topic = os.path.splitext(docx_file.filename)[0]
    docx_path = os.path.join(UPLOAD_FOLDER, docx_file.filename)
    docx_file.save(docx_path)

    doc = Document(docx_path)
    questions = []
    current_q = ""
    options = []

    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue

        # Detect new question: starts with number or ends with ?
        if re.match(r'^\d+\.', text) or text.endswith("?"):
            if current_q:
                if options:
                    # Save question + options (join options with semicolon!)
                    current_q += " Options: " + "; ".join(options)
                questions.append(current_q.strip())
            current_q = text
            options = []
        else:
            # Assume this is an option (bullet point or indented)
            options.append(text)

    # Final question
    if current_q:
        if options:
            current_q += " Options: " + "; ".join(options)
        questions.append(current_q.strip())

    user_responses = []
    explanation_cache = {}
    return jsonify({"message": "Files uploaded", "count": len(questions)})

@app.route('/get_questions')
def get_questions():
    return jsonify({"questions": questions})
word_to_number = {
    "zero": "0", "one": "1", "two": "2", "three": "3", "four": "4",
    "five": "5", "six": "6", "seven": "7", "eight": "8", "nine": "9",
    "ten": "10", "eleven": "11", "twelve": "12", "thirteen": "13",
    "fourteen": "14", "fifteen": "15", "sixteen": "16", "seventeen": "17",
    "eighteen": "18", "nineteen": "19", "twenty": "20", "thirty": "30",
    "forty": "40", "fifty": "50", "sixty": "60", "seventy": "70",
    "eighty": "80", "ninety": "90", "hundred": "100", "thousand": "000"
}

def word_to_num(text):
    # Replace word-numbers with digits where appropriate
    for word, num in word_to_number.items():
        text = re.sub(rf'\b{word}\b', num, text, flags=re.IGNORECASE)
    return text

yes_variants = ["yes", "ya", "yeah", "yup", "sure", "interested"]
no_variants = ["no", "nope", "nah", "not interested"]

def normalize_user_input(text):
    text = text.lower().strip()
    text = word_to_num(text)
    
    if any(word in text for word in yes_variants):
        return "yes"
    if any(word in text for word in no_variants):
        return "no"

    # existing fixes
    text = re.sub(r"\b(\d+)\s*(to|too|two)\s*(\d+)\b", r"\1 to \3", text)
    text = re.sub(r"\bstate\b", "stick", text)
    return text


@app.route('/submit', methods=['POST'])
def submit():
    global user_responses
    data = request.get_json()
    question = data['question']
    user_input_raw = data['answer']
    user_input = normalize_user_input(user_input_raw)

    matched_option = None
    options_text = ""

    if "Options:" in question:
        parts = question.split("Options:")
        question_text = parts[0].strip()
        options_text = parts[1].strip()

        options_raw = [opt.strip() for opt in re.split(r"[;]", options_text)] # original case
        options_list = [opt.lower() for opt in options_raw]             # lowercased for matching
        option_map = dict(zip(options_list, options_raw))               # map lower → original

        # 🔍 Ask Ollama to choose best match
        prompt = f"""
            You are a smart survey assistant helping categorize user responses.

            🎯 TASK:
            From the list of options, choose the **best matching option** that aligns with the user’s answer.
            Only respond with:
            - An **exact option** from the list (no explanation)
            - Or respond with **"other"** if none are relevant
            - Or respond with **"retry"** if the answer is off-topic, gibberish, or unclear.

            ✅ Example:
            Question: What are they using to support their mobility?
            Options: stick, walker, none
            User: They are using a stick for walking.
            Your answer: stick

            🧾 QUESTION: {question_text}
            📋 OPTIONS: {', '.join(options_raw)}
            🗣️ USER SAID: {user_input}

            Your reply (only one word: exact option, or 'other', or 'retry'):
        """
        model_reply = get_ai_reply(prompt).lower().strip()
        model_reply = re.sub(r"[^a-zA-Z0-9 ]", "", model_reply).strip()

        # ✅ Direct match shortcut
        if model_reply in option_map:
            matched_option = option_map[model_reply]
        else:
            # 🔁 Fuzzy matching
            best_match = None
            highest_score = 0
            for opt in options_list:
                score = max(
                    fuzz.ratio(model_reply, opt),
                    fuzz.partial_ratio(model_reply, opt),
                    fuzz.token_sort_ratio(model_reply, opt)
                )
                if model_reply in opt:
                    score += 10  # bonus for containment

                if score > highest_score:
                    highest_score = score
                    best_match = opt

            if highest_score >= 85:
                matched_option = option_map.get(best_match, best_match)
            elif model_reply == "other":
                matched_option = "other"
            elif model_reply == "retry":
                return jsonify({
                    "reply": f"Sorry, I couldn't understand that. Please choose from the following options: {options_text}",
                    "retry": True
                })
            else:
                matched_option = "other"

        # 🔁 Optional fallback: if cleaned user_input matches any option
        if not matched_option and user_input in option_map:
            matched_option = option_map[user_input]

    ai_reply = f"Got it. You've selected '{matched_option}'. Thanks!"
    user_responses.append((question, matched_option, ai_reply, user_input_raw))

    return jsonify({
        "reply": ai_reply,
        "retry": False
    })

@app.route("/finalize", methods=["POST"])
def finalize():
    global survey_topic, questions, user_responses
    if not survey_topic:
        return jsonify({"error": "No topic set."}), 400

    file_path = os.path.join(UPLOAD_FOLDER, f"{survey_topic}.xlsx")

    try:
        if os.path.exists(file_path):
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            col = 1
            for q in questions:
                ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
                ws.cell(row=1, column=col).value = q
                ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')
                ws.cell(row=2, column=col).value = "User Answer"
                ws.cell(row=2, column=col+1).value = "AI Reply"
                col += 2

        row = ws.max_row + 1
        col = 1
        for i, (question, user_ans, ai_ans, _) in enumerate(user_responses):
            ws.cell(row=row, column=col).value = user_ans
            ws.cell(row=row, column=col+1).value = ai_ans
            col += 2

        wb.save(file_path)
        print("✅ Saved to", file_path)
        return jsonify({"message": "Responses saved."})
    except PermissionError:
        return jsonify({"error": "❌ Please close the Excel file before saving!"}), 500
    except Exception as e:
        print("Error saving:", e)
        return jsonify({"error": "An error occurred while saving the file."}), 500


@app.route("/explain", methods=["POST"])
def explain():
    data = request.get_json()
    question = data["question"]

    # Use Ollama AI model to generate an explanation
    prompt = (
        f"You are a friendly survey assistant. "
        f"Give a very short, clear explanation of the question so the user can understand it better. "
        f"Do NOT repeat the original question. "
        f"Do NOT mention or list the options in the explanation.\n\n"
        f"Question: {question}\n"
        f"Explanation:"
    )

    try:
        response = requests.post(
            "http://localhost:11434/api/generate",
            json={
                "model": "mistral",  # or your preferred model like llama3 or gemma
                "prompt": prompt,
                "stream": False
            }
        )
        explanation = response.json()["response"].strip()
        return jsonify({"explanation": explanation})
    except Exception as e:
        return jsonify({"explanation": "Sorry, I couldn't explain the question right now."})

if __name__ == '__main__':
    app.run(debug=True)
